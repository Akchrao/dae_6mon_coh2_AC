import openpyxl, pickle, copy
from datetime import datetime

with open("styles.pkl","rb") as f:
    styles = pickle.load(f)

PERSON = "Akhila Chennamaneni"

weeks = [
{
 "title": "Week 1 - Environment & Identity Provider Setup",
 "dates": ["2026-08-17","2026-08-18","2026-08-19","2026-08-20"],
 "subitem": "Weekly Status Update",
 "days": [
  {"title":"Day 1: Define Project Scope, Use Case & Success Criteria",
   "desc":"Write the problem statement, in-scope and out-of-scope boundaries, the 5 healthcare job roles, and 3-4 measurable success criteria for the project.",
   "deliv":"Scope document (1-2 pages) + role list + success criteria checklist"},
  {"title":"Day 2: Set Up Project Repository & Local Environment",
   "desc":"Create the GitHub repository with a docs/scripts/configs/data/outputs folder structure. Install Python, VS Code, and Git. Create a virtual environment and install core packages (pandas, flask, faker).",
   "deliv":"GitHub repo live with folder structure, README, and requirements.txt"},
  {"title":"Day 3: Install & Configure Identity Provider (Keycloak or Entra ID)",
   "desc":"Install Keycloak locally, or set up a Microsoft Entra ID free tenant. Create a realm/directory named CareAccess for the fictional organization and configure the base admin settings.",
   "deliv":"Running Keycloak instance (or Entra tenant) with a CareAccess realm created"},
  {"title":"Day 4: Set Up Database & Validate Full Environment",
   "desc":"Install and configure SQLite (or PostgreSQL). Create the initial schema tables for users, roles, departments, and applications. Write a Python script that connects to the database, and confirm every tool from Days 1-3 works together.",
   "deliv":"database.db with initial schema + passing connection test script + environment checklist signed off"},
 ]},
{
 "title": "Week 2 - Identity Data Model & Synthetic Organization",
 "dates": ["2026-08-24","2026-08-25","2026-08-26","2026-08-27"],
 "subitem": "Weekly Status Update",
 "days": [
  {"title":"Day 1: Design the Identity Data Model",
   "desc":"Design the entity-relationship diagram covering users, departments, roles, applications, and entitlements, and define the attributes each entity will hold.",
   "deliv":"ER diagram (Draw.io) + schema.sql"},
  {"title":"Day 2: Generate the Synthetic Healthcare Workforce",
   "desc":"Use Faker to generate 30 synthetic users across the 5 job roles (Physician, Nurse, Admin Staff, Lab Technician, IT Support) with realistic attributes and department assignments. No real names or PHI.",
   "deliv":"users.csv containing 30 synthetic identities"},
  {"title":"Day 3: Load Users Into the Identity Provider",
   "desc":"Import the 30 synthetic users into Keycloak (or Entra ID), assign each to their department group, and verify every account is queryable through the admin console or API.",
   "deliv":"30 verified user accounts live in the identity provider"},
  {"title":"Day 4: Define Applications & the Entitlement Catalog",
   "desc":"Define the mock EHR application(s) and the full list of permissions each one exposes (e.g., view patient record, edit orders, billing access). This becomes the basis for RBAC.",
   "deliv":"applications.csv + entitlements.csv, plus a short note confirming Weeks 1-2 environment works end to end"},
 ]},
{
 "title": "Week 3 - RBAC Design & Enforcement",
 "dates": ["2026-08-31","2026-09-01","2026-09-02","2026-09-03"],
 "subitem": "Weekly Status Update",
 "days": [
  {"title":"Day 1: Design the RBAC Role-to-Permission Matrix",
   "desc":"Map each of the 5 job roles to the entitlements it should have in the entitlement catalog, applying least-privilege principles.",
   "deliv":"Draft RBAC matrix (roles x permissions)"},
  {"title":"Day 2: Build the Mock EHR Application Skeleton",
   "desc":"Build a basic Flask app representing the EHR system with CSV-backed tables for patients, visits, and orders, wired to check a user's role before returning data.",
   "deliv":"Flask app skeleton + patients.csv, visits.csv, orders.csv"},
  {"title":"Day 3: Implement RBAC Enforcement Logic",
   "desc":"Write the enforcement code that checks a user's role against the RBAC matrix before allowing an action in the mock EHR app.",
   "deliv":"Working enforcement function integrated into the Flask app"},
  {"title":"Day 4: Test RBAC Enforcement Across All 5 Roles",
   "desc":"Run each of the 5 roles against every permission type and confirm access is correctly allowed or denied. Log and fix any mismatches.",
   "deliv":"RBAC test log showing pass/fail results for all role-permission combinations"},
 ]},
{
 "title": "Week 4 - Governance Foundations",
 "dates": ["2026-09-07","2026-09-08","2026-09-09","2026-09-10"],
 "subitem": "Monthly Status Update / End of Semester",
 "days": [
  {"title":"Day 1: Finalize the Access Control Matrix",
   "desc":"Incorporate Week 3's test results into a finalized, clean version of the RBAC/access control matrix.",
   "deliv":"Final RBAC Access Matrix.xlsx"},
  {"title":"Day 2: Write Access Policies & Provisioning SOPs",
   "desc":"Write the standard operating procedures covering how access is requested, approved, provisioned, and removed.",
   "deliv":"Access Policy & SOP document"},
  {"title":"Day 3: Build the Risk Register & Break-Glass Procedure",
   "desc":"Create a risk register listing the top access-related risks (dormant accounts, excessive privilege, terminated-user access) with likelihood/impact ratings. Write a break-glass procedure for emergency access during a system outage.",
   "deliv":"Risk Register.xlsx + Break-Glass Procedure document"},
  {"title":"Day 4: Review & Sign Off on Governance Documentation",
   "desc":"Walk through the RBAC matrix, SOPs, risk register, and break-glass procedure end to end and confirm they are internally consistent before moving into lifecycle automation.",
   "deliv":"Signed-off governance package"},
 ]},
{
 "title": "Week 5 - Joiner & Mover Automation",
 "dates": ["2026-09-14","2026-09-15","2026-09-16","2026-09-17"],
 "subitem": "Weekly Status Update",
 "days": [
  {"title":"Day 1: Build the Joiner Workflow",
   "desc":"Write the onboarding script: create the user record, assign the correct role/department, and provision the entitlements defined in the RBAC matrix.",
   "deliv":"joiner.py script"},
  {"title":"Day 2: Test the Joiner Workflow",
   "desc":"Run 5 test onboarding scenarios (one per role) through the joiner script and confirm each new hire receives exactly the access their role should have.",
   "deliv":"Joiner test log + 5 verified onboarding records"},
  {"title":"Day 3: Build the Mover Workflow",
   "desc":"Write the role/department-change script: revoke entitlements tied to the old role and grant the entitlements tied to the new role.",
   "deliv":"mover.py script"},
  {"title":"Day 4: Test the Mover Workflow",
   "desc":"Run test role-change scenarios and confirm old access is fully revoked and new access is fully granted with no orphaned entitlements.",
   "deliv":"Mover test log"},
 ]},
{
 "title": "Week 6 - Leaver Automation & JML Integration",
 "dates": ["2026-09-21","2026-09-22","2026-09-23","2026-09-24"],
 "subitem": "Weekly Status Update",
 "days": [
  {"title":"Day 1: Build the Leaver Workflow",
   "desc":"Write the termination script: revoke all access, archive the user record, and generate an audit record of what was removed.",
   "deliv":"leaver.py script"},
  {"title":"Day 2: Test the Leaver Workflow",
   "desc":"Run test termination scenarios and confirm zero residual access remains for terminated users.",
   "deliv":"Leaver test log confirming zero residual access"},
  {"title":"Day 3: Integrate Joiner, Mover & Leaver Into One Pipeline",
   "desc":"Combine the three scripts into a single JML command-line tool that can process onboarding, transfer, and termination events from one entry point.",
   "deliv":"jml_pipeline.py (unified CLI)"},
  {"title":"Day 4: Generate & Validate JML Audit Logs",
   "desc":"Run the full pipeline across a batch of test events and export the resulting audit trail to CSV. Confirm every event is logged correctly.",
   "deliv":"jml_audit_log.csv"},
 ]},
{
 "title": "Week 7 - Access Request & Approval",
 "dates": ["2026-09-28","2026-09-29","2026-09-30","2026-10-01"],
 "subitem": "Weekly Status Update",
 "days": [
  {"title":"Day 1: Build the Access Request Workflow",
   "desc":"Build a mock access-request form and the backend logic that captures what a user is requesting and why.",
   "deliv":"Access request module + sample requests"},
  {"title":"Day 2: Build the Approval Workflow",
   "desc":"Build approval routing logic based on role: manager approval, plus IT/compliance approval for privileged access.",
   "deliv":"Approval routing module"},
  {"title":"Day 3: Integrate Request, Approval & Automatic Provisioning",
   "desc":"Connect the request and approval modules so that an approved request automatically triggers provisioning through the JML pipeline.",
   "deliv":"End-to-end request-to-provisioning flow"},
  {"title":"Day 4: Test the Full Access Request Flow",
   "desc":"Submit test requests through the entire flow (request, approval, provisioning) and confirm access is granted correctly and only after approval.",
   "deliv":"End-to-end test log"},
 ]},
{
 "title": "Week 8 - Access Review & Compliance",
 "dates": ["2026-10-05","2026-10-06","2026-10-07","2026-10-08"],
 "subitem": "Monthly Status Update / End of Semester",
 "days": [
  {"title":"Day 1: Build the Access Review / Certification Process",
   "desc":"Build a periodic access-review campaign where each manager attests to whether their team's access is still appropriate.",
   "deliv":"Access review campaign module"},
  {"title":"Day 2: Implement Compliance Checks & Risk Flags",
   "desc":"Add automated checks that flag dormant accounts, expired contractors, role mismatches, and privileged access outside policy.",
   "deliv":"Compliance rule set + flagged-user output"},
  {"title":"Day 3: Generate Governance Reports",
   "desc":"Build reporting tables summarizing review outcomes, flagged users, and overall compliance status.",
   "deliv":"Governance report (CSV + summary table)"},
  {"title":"Day 4: Test & Validate Governance Features End to End",
   "desc":"Run a full certification campaign, verify the risk flags are accurate, and fix any false positives found.",
   "deliv":"Validated governance module"},
 ]},
{
 "title": "Week 9 - Monitoring & Incident Simulation",
 "dates": ["2026-10-12","2026-10-13","2026-10-14","2026-10-15"],
 "subitem": "Weekly Status Update",
 "days": [
  {"title":"Day 1: Build the Security Event Log Generator",
   "desc":"Write a script that simulates 30 days of login, access, and approval activity across the synthetic organization.",
   "deliv":"activity_log.csv (30 days of simulated events)"},
  {"title":"Day 2: Simulate 3 Security Incidents",
   "desc":"Inject three incident scenarios into the log data: a terminated user attempting access, excessive permission usage, and a suspicious access pattern.",
   "deliv":"3 labeled incident scenarios embedded in the log data"},
  {"title":"Day 3: Analyze Logs & Detect Anomalies",
   "desc":"Use Python/pandas to correlate the log data and identify the 3 injected incidents. Document what was found and how it was detected.",
   "deliv":"Incident analysis notebook/report"},
  {"title":"Day 4: Test the Break-Glass Emergency Access Procedure",
   "desc":"Simulate a system outage scenario and walk through the break-glass procedure defined in Week 4 to confirm it works in practice.",
   "deliv":"Break-glass test log"},
 ]},
{
 "title": "Week 10 - Dashboards & Stabilization",
 "dates": ["2026-10-19","2026-10-20","2026-10-21","2026-10-22"],
 "subitem": "Weekly Status Update",
 "days": [
  {"title":"Day 1: Connect Data to Tableau & Build an Initial Dashboard",
   "desc":"Connect the CSV outputs to Tableau Public and build the first pass of visuals for access and incident data.",
   "deliv":"Initial Tableau workbook"},
  {"title":"Day 2: Build the Full Governance & Incident Dashboard",
   "desc":"Expand the dashboard to cover access metrics, review/compliance status, and incident trends in one view.",
   "deliv":"Completed Tableau dashboard"},
  {"title":"Day 3: Stabilize the Project - Fix Defects Across All Modules",
   "desc":"Run a full bug bash across the JML pipeline, access request flow, and governance modules. Fix anything broken.",
   "deliv":"Defect log with fixes applied"},
  {"title":"Day 4: Clean & Validate All CSV Outputs",
   "desc":"Check every CSV output produced so far for consistency, completeness, and correct formatting.",
   "deliv":"Validated final dataset folder"},
 ]},
{
 "title": "Week 11 - Documentation & Packaging",
 "dates": ["2026-10-26","2026-10-27","2026-10-28","2026-10-29"],
 "subitem": "Weekly Status Update",
 "days": [
  {"title":"Day 1: Complete Technical Documentation",
   "desc":"Write the setup guide, architecture overview, and data dictionary covering every component built.",
   "deliv":"Technical documentation package"},
  {"title":"Day 2: Complete User Documentation",
   "desc":"Write the SOPs and runbook explaining how each workflow (JML, access request, review) operates in plain language.",
   "deliv":"User documentation / runbook"},
  {"title":"Day 3: Package the Complete IAM Solution",
   "desc":"Bundle all scripts, configs, datasets, and documentation into a single organized release package.",
   "deliv":"Release package (zip) with README"},
  {"title":"Day 4: Deploy & Run Final Acceptance Testing",
   "desc":"Deploy the packaged solution to a clean demonstration environment and run final acceptance tests against the success criteria defined in Week 1.",
   "deliv":"UAT results confirming all success criteria met"},
 ]},
{
 "title": "Week 12 - Demonstration & Wrap-Up",
 "dates": ["2026-11-02","2026-11-03","2026-11-04","2026-11-05"],
 "subitem": None,
 "days": [
  {"title":"Day 1: Prepare Final Dashboard Outputs, Screenshots & Diagrams",
   "desc":"Capture clean screenshots of every workflow, the dashboard, and sample reports for the final presentation.",
   "deliv":"Screenshot and diagram set for the demo"},
  {"title":"Day 2: Build the Demonstration Materials",
   "desc":"Build the presentation deck and demo script walking through the full IAM solution end to end.",
   "deliv":"Demo deck + demo script"},
  {"title":"Day 3: Present the Project Demonstration",
   "desc":"Deliver the live (or recorded) demo, walking through the RBAC model, JML automation, access governance, and monitoring dashboard.",
   "deliv":"Completed demo + feedback notes"},
  {"title":"Day 4: Document Lessons Learned & Future Roadmap",
   "desc":"Summarize what worked, what was hard, and what you'd change. Propose 2-3 future enhancements (e.g., SSO integration, self-service portal, expanded roles).",
   "deliv":"Lessons Learned & Roadmap document"},
 ]},
]

# Generate sequential item IDs
item_id_counter = 20000000001
def next_id():
    global item_id_counter
    item_id_counter += 7
    return str(item_id_counter)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "akhila's project board"
ws2 = wb.create_sheet("updates")

def setcell(sheet, r, c, value, style_key):
    cell = sheet.cell(row=r, column=c, value=value)
    cell._style = copy.copy(styles[style_key])
    return cell

ws.column_dimensions['A'].width = 40
ws.column_dimensions['B'].width = 20

row = 1
setcell(ws, row, 1, "Akhila's Project Board", 'title')
row += 1

update_rows = []  # (item_id, item_name, created_at, content)
created_counter = 0

for week in weeks:
    setcell(ws, row, 1, week["title"], 'week_header')
    row += 1
    headers = ['Name','Subitems','Status','Person','Date','Timeline - Start','Timeline - End','Text','Checkbox','Item ID (auto generated)']
    for c, h in enumerate(headers, start=1):
        setcell(ws, row, c, h, 'colheader')
    row += 1

    day_item_ids = []
    for i, day in enumerate(week["days"]):
        item_id = next_id()
        day_item_ids.append(item_id)
        setcell(ws, row, 1, day["title"], 'day_name')
        if i == 3 and week["subitem"]:
            setcell(ws, row, 2, week["subitem"], 'day_subitem')
        setcell(ws, row, 3, None, 'day_status')
        setcell(ws, row, 4, PERSON, 'day_person')
        d = datetime.strptime(week["dates"][i], "%Y-%m-%d")
        setcell(ws, row, 5, d, 'day_date')
        setcell(ws, row, 10, item_id, 'day_itemid')
        row += 1

        created_counter += 1
        created_at = f"06/August/2026  0{1+created_counter//60:01d}:{(created_counter*3)%60:02d}:00 AM"
        content = f"Description: {day['desc']}\n\nDeliverable: {day['deliv']}"
        update_rows.append((item_id, day["title"], created_at, content))

    if week["subitem"]:
        setcell(ws, row, 1, "Subitems", 'subhdr_label')
        subhdrs = ['Subitems','Name','Owner','Status','Date','Item ID (auto generated)']
        for c, h in enumerate(subhdrs[1:], start=2):
            setcell(ws, row, c, h, 'subhdr_col')
        row += 1
        sub_id = next_id()
        setcell(ws, row, 2, week["subitem"], 'subitem_name')
        setcell(ws, row, 4, None, 'subitem_status')
        setcell(ws, row, 6, sub_id, 'subitem_itemid')
        row += 1
        created_counter += 1
        created_at = f"06/August/2026  0{1+created_counter//60:01d}:{(created_counter*3)%60:02d}:00 AM"
        content = f"Status checkpoint covering all four sessions this week: {week['title']}."
        update_rows.append((sub_id, week["subitem"], created_at, content))

    setcell(ws, row, 5, f"{week['dates'][0]} to {week['dates'][3]}", 'summary_date')
    setcell(ws, row, 6, None, 'summary_blank')
    setcell(ws, row, 7, None, 'summary_blank')
    setcell(ws, row, 9, "0/4", 'summary_count')
    row += 1
    row += 1  # blank spacer row

# Updates sheet
setcell(ws2, 1, 1, "Akhila's Project Board", 'title')
ws2.cell(row=1, column=2, value="Updates")
update_headers = ['Item ID','Item Name','Content Type','Content Type','User','Created At','Update Content','Likes Count','Asset IDs','Post ID','Parent Post ID']
for c, h in enumerate(update_headers, start=1):
    cell = ws2.cell(row=2, column=c, value=h)
    cell._style = copy.copy(styles['colheader'])

post_id_counter = 6000000001
r = 3
for item_id, item_name, created_at, content in update_rows:
    post_id_counter += 11
    ws2.cell(row=r, column=1, value=item_id)
    ws2.cell(row=r, column=2, value=item_name)
    ws2.cell(row=r, column=3, value="Update")
    ws2.cell(row=r, column=5, value=PERSON)
    ws2.cell(row=r, column=6, value=created_at)
    ws2.cell(row=r, column=7, value=content)
    ws2.cell(row=r, column=8, value=0)
    ws2.cell(row=r, column=10, value=str(post_id_counter))
    r += 1

ws2.column_dimensions['A'].width = 15
ws2.column_dimensions['B'].width = 45
ws2.column_dimensions['G'].width = 80
for col in ['C','D','E','F','H','I','J','K']:
    ws2.column_dimensions[col].width = 16

wb.save("/mnt/user-data/outputs/Akhila_Project_Board_Restructured.xlsx")
print("Saved. Total days:", sum(len(w['days']) for w in weeks))
